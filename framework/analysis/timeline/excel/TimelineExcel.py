import json
from openpyxl import Workbook
from framework.analysis.Export import Export


class TimelineExcel(Export):

    def to_file(self, output_path, result_json):
        result = json.loads(result_json)
        wb = Workbook()
        ws = wb.active
        ws['A1'] = result['title']
        ws['B1'] = 'Created at:'
        ws['C1'] = result['created_at']

        for element in result['elements']:
            ws.append([
                element['datetime'],
                element['title'],
                element['content']
            ])

        wb.save(output_path)
