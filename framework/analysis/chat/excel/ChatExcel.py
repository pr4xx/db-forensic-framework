import json
from openpyxl import Workbook
from framework.analysis.Export import Export


class ChatExcel(Export):

    def to_file(self, output_path, result_json):
        result = json.loads(result_json)
        wb = Workbook()
        ws = wb.active
        ws.title = 'Overview'
        ws['A1'] = result['title']
        ws['B1'] = 'Created at:'
        ws['C1'] = result['created_at']

        index = 0
        for conversation in result['conversations']:
            index += 1
            new_ws = wb.create_sheet(title="Conversation " + str(index))
            for message in conversation['messages']:
                new_ws.append([
                    message['participant']['name'],
                    message['datetime'],
                    message['content']
                ])

        wb.save(output_path)
