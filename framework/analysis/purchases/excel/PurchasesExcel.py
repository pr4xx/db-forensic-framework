import json
from openpyxl import Workbook
from framework.analysis.Export import Export


class PurchasesExcel(Export):

    def to_file(self, output_path, result_json):
        result = json.loads(result_json)
        wb = Workbook()
        ws = wb.active
        ws.title = 'Overview'
        ws['A1'] = result['title']
        ws['B1'] = 'Created at:'
        ws['C1'] = result['created_at']

        index = 0
        for card in result['shopping_cards']:
            index += 1
            new_ws = wb.create_sheet(title="Card " + str(index))
            for article in card['articles']:
                new_ws['A1'] = 'Invoice Address:'
                new_ws['B1'] = card['invoice_address']
                new_ws['C1'] = 'Shipping Address:'
                new_ws['D1'] = card['shipping_address']
                new_ws['E1'] = 'Shipping:'
                new_ws['F1'] = card['shipping']
                new_ws['A2'] = 'Product'
                new_ws['B2'] = 'Description'
                new_ws['C2'] = 'Amount'
                new_ws['D2'] = 'Total'
                new_ws.append([
                    article['name'],
                    article['description'],
                    article['amount'],
                    article['total']
                ])

        wb.save(output_path)
